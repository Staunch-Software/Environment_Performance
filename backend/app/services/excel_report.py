"""Excel report generation with 5 sheets."""
import uuid
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.orb_upload import OrbUpload
from app.models.orb_entry import OrbEntry
from app.models.orb_entry_quantity import OrbEntryQuantity
from app.models.orb_alert import OrbAlert
from app.models.vessel import Vessel
from app.models.vessel_tank import VesselTank


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
SEVERITY_FILLS = {
    "critical": PatternFill("solid", fgColor="FFE0E0"),
    "major": PatternFill("solid", fgColor="FFF0E0"),
    "minor": PatternFill("solid", fgColor="FFFDE0"),
    "observation": PatternFill("solid", fgColor="F5F5F5"),
}
ORANGE_FILL = PatternFill("solid", fgColor="FFA500")


def auto_fit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


async def generate_excel(upload_id: uuid.UUID, db: AsyncSession) -> bytes:
    upload_result = await db.execute(select(OrbUpload).where(OrbUpload.id == upload_id))
    upload = upload_result.scalar_one_or_none()

    vessel_result = await db.execute(select(Vessel).where(Vessel.id == upload.vessel_id))
    vessel = vessel_result.scalar_one_or_none()

    entries_result = await db.execute(
        select(OrbEntry).where(OrbEntry.upload_id == upload_id).order_by(OrbEntry.entry_date)
    )
    entries = entries_result.scalars().all()

    alerts_result = await db.execute(
        select(OrbAlert).where(OrbAlert.vessel_id == upload.vessel_id).order_by(OrbAlert.created_at.desc())
    )
    alerts = alerts_result.scalars().all()

    tanks_result = await db.execute(
        select(VesselTank).where(VesselTank.vessel_id == upload.vessel_id, VesselTank.is_active == True)
    )
    tanks = tanks_result.scalars().all()

    entry_quantities: dict[uuid.UUID, list] = {}
    for entry in entries:
        qty_result = await db.execute(
            select(OrbEntryQuantity).where(OrbEntryQuantity.entry_id == entry.id)
        )
        entry_quantities[entry.id] = qty_result.scalars().all()

    min_date = min((e.entry_date for e in entries), default=None)
    max_date = max((e.entry_date for e in entries), default=None)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: ORB Entries ──────────────────────────────────────────────────
    ws1 = wb.create_sheet("ORB Entries")
    ws1.append(["OIL RECORD BOOK – PART I (Machinery Space Operations)"])
    ws1.append([f"Vessel Name: {vessel.name if vessel else 'N/A'}"])
    ws1.append([f"IMO Number: {vessel.imo_number if vessel else 'N/A'}"])
    ws1.append([f"Period covered: {min_date} to {max_date}"])
    ws1.append([])

    headers = [
        "Sr.No.", "Date (dd-mmm-yyyy)", "Code", "Item",
        "Operation / Particulars", "Qty (m³/MT)", "Tank / Location",
        "Time / Position", "Officer 1", "Officer 2",
    ]
    ws1.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=6, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)

    for idx, entry in enumerate(entries, 1):
        quantities = entry_quantities.get(entry.id, [])
        qty_str = " / ".join(
            f"{q.qty_value} {q.qty_unit} ({q.qty_type})" for q in quantities
        )
        officer1 = f"{entry.officer_1_name or ''} {entry.officer_1_rank or ''}".strip()
        officer2 = f"{entry.officer_2_name or ''} {entry.officer_2_rank or ''}".strip()
        time_pos = " / ".join(filter(None, [entry.time_start, entry.position_start]))

        row_data = [
            idx,
            str(entry.entry_date),
            entry.orb_code,
            entry.item_number or "",
            entry.operation_description,
            qty_str,
            entry.tank_location or "",
            time_pos,
            officer1,
            officer2,
        ]
        ws1.append(row_data)
        fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
        for col_num in range(1, len(headers) + 1):
            ws1.cell(row=6 + idx, column=col_num).fill = fill

    auto_fit(ws1)

    # ── Sheet 2: Discrepancy Register ─────────────────────────────────────────
    ws2 = wb.create_sheet("Discrepancy Register")
    headers2 = ["Sr.No.", "Severity", "Alert Type", "Message", "Entry Date", "Created At", "Status"]
    ws2.append(headers2)
    for col_num, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for idx, alert in enumerate(alerts, 1):
        row = [
            idx,
            alert.severity,
            alert.alert_type,
            alert.message,
            str(alert.created_at.date()) if alert.created_at else "",
            str(alert.created_at) if alert.created_at else "",
            "Resolved" if alert.is_resolved else "Open",
        ]
        ws2.append(row)
        fill = SEVERITY_FILLS.get(alert.severity, WHITE_FILL)
        for col_num in range(1, len(headers2) + 1):
            ws2.cell(row=idx + 1, column=col_num).fill = fill

    auto_fit(ws2)

    # ── Sheet 3: Tank Running Balance ─────────────────────────────────────────
    ws3 = wb.create_sheet("Tank Running Balance")
    headers3 = ["Date", "Tank", "Opening (m³)", "In (m³)", "Out (m³)",
                "Computed Closing (m³)", "Logged Closing (m³)", "Δ (m³)"]
    ws3.append(headers3)
    for col_num, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_num = 2
    current_group = None
    for tank in sorted(tanks, key=lambda t: (t.tank_group or "Ungrouped", t.tank_name)):  # ← sort by group
        # Insert a group header row when group changes
        if tank.tank_group != current_group:
            current_group = tank.tank_group or "Ungrouped"
            ws3.append([current_group])   # ← group label row
            for col_num in range(1, len(headers3) + 1):
                cell = ws3.cell(row=row_num, column=col_num)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            row_num += 1

        balance = None
        for entry in entries:
            quantities = entry_quantities.get(entry.id, [])
            inflow = outflow = logged_closing = None

            for q in quantities:
                if q.qty_type == "retained" and (q.from_tank == tank.tank_name or q.to_tank == tank.tank_name):
                    logged_closing = q.qty_value
                elif q.qty_type == "transferred":
                    if q.to_tank == tank.tank_name:
                        inflow = (inflow or 0) + q.qty_value
                    elif q.from_tank == tank.tank_name:
                        outflow = (outflow or 0) + q.qty_value
                elif q.qty_type in ("disposed", "evaporated") and q.from_tank == tank.tank_name:
                    outflow = (outflow or 0) + q.qty_value

            if inflow is None and outflow is None and logged_closing is None:
                continue

            opening = balance
            computed = (opening or 0) + (inflow or 0) - (outflow or 0)
            delta = abs(computed - logged_closing) if logged_closing is not None else None

            row_data = [
                str(entry.entry_date),
                tank.tank_name,
                round(opening, 3) if opening is not None else "",
                round(inflow, 3) if inflow is not None else 0,
                round(outflow, 3) if outflow is not None else 0,
                round(computed, 3),
                round(logged_closing, 3) if logged_closing is not None else "",
                round(delta, 3) if delta is not None else "",
            ]
            ws3.append(row_data)

            if delta is not None and delta > 0.15:
                for col_num in range(1, len(headers3) + 1):
                    ws3.cell(row=row_num, column=col_num).fill = ORANGE_FILL

            if logged_closing is not None:
                balance = logged_closing
            row_num += 1

    auto_fit(ws3)

    # ── Sheet 4: Bunkering Summary ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Bunkering Summary")
    headers4 = ["Date", "Place / Port", "Grade", "Quantity (MT)",
                "Sulphur %", "Density (kg/L)", "Viscosity", "Tanks Loaded", "BDN Ref"]
    ws4.append(headers4)
    for col_num, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    bunker_entries = [e for e in entries if e.orb_code == "H" and e.item_number == "26.3"]
    total_bunker = 0.0
    for idx, entry in enumerate(bunker_entries, 1):
        quantities = entry_quantities.get(entry.id, [])
        qty_val = sum(q.qty_value for q in quantities if q.qty_type == "bunkered")
        total_bunker += qty_val
        fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
        row = [
            str(entry.entry_date), "", "", round(qty_val, 3),
            "", "", "", entry.tank_location or "", "",
        ]
        ws4.append(row)
        for col_num in range(1, len(headers4) + 1):
            ws4.cell(row=idx + 1, column=col_num).fill = fill

    ws4.append(["TOTAL", "", "", round(total_bunker, 3), "", "", "", "", ""])
    total_row = ws4.max_row
    for col_num in range(1, len(headers4) + 1):
        cell = ws4.cell(row=total_row, column=col_num)
        cell.font = Font(bold=True)

    auto_fit(ws4)

    # ── Sheet 5: Methodology & Accuracy ───────────────────────────────────────
    ws5 = wb.create_sheet("Methodology & Accuracy")
    content = [
        ["ORB DIGITIZATION — METHODOLOGY & ACCURACY STATEMENT"],
        [],
        ["1. EXTRACTION METHOD"],
        ["This report was generated by the ORB Digitization Platform using AI-assisted OCR extraction."],
        ["Handwritten ORB pages were processed using Claude AI (claude-sonnet-4-6) vision capabilities."],
        [],
        ["2. CONFIDENCE SCORING"],
        ["Each extracted entry carries a confidence score from 0.0 (unreadable) to 1.0 (perfectly legible)."],
        ["Entries with confidence < 0.75 are flagged for manual review and highlighted in the platform."],
        [],
        ["3. REGULATORY REFERENCES"],
        ["MARPOL Annex I, Regulation 17 — Oil Record Book Part I"],
        ["MEPC.1/Circ.736/Rev.3 — Guidance for the recording of operations in the ORB"],
        ["Resolution MEPC.117(52) — Amendments to MARPOL Annex I"],
        [],
        ["4. COMPLIANCE CHECKS PERFORMED"],
        ["- Running balance verification per tank (±0.15 m³ tolerance)"],
        ["- Individual tank capacity threshold (85%)"],
        ["- Combined tank capacity threshold (85%)"],
        ["- Overdue sounding detection (>8 day gaps)"],
        ["- MARPOL code violation flags (Items 12.4, 11.4, Code I)"],
        ["- Overdue bilge discharge detection (>14 days)"],
        ["- Missing BDN reference in bunkering entries"],
        ["- Sludge generation rate (0.5%–2% of fuel bunkered)"],
        ["- Low confidence extraction flagging"],
        [],
        ["5. DISCLAIMER"],
        ["This report is generated from digitized handwritten records and may contain transcription errors."],
        ["All flagged entries MUST be verified against original ORB documents before any regulatory submission."],
        ["This tool does not replace the legal obligation to maintain accurate ORB records."],
        ["The vessel master and chief engineer remain responsible for ORB accuracy under MARPOL."],
    ]
    for row in content:
        ws5.append(row)
    ws5.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
